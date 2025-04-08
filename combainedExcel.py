# from sampleJson import formattedSheet
import json
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Font, Side
from openpyxl.worksheet.pagebreak import Break
apply_border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")                 
    )
def increment_column(column):
    # If column is Z, return AA
    if column == 'Z':
        return 'AA'
    # If column has multiple characters
    if len(column) == 1:
        next_char = chr(ord(column) + 1)
        return next_char
    else:
        last_char = column[-1]
        if last_char == 'Z':
            return increment_column(column[:-1]) + 'A'
        else:
            return column[:-1] + chr(ord(last_char) + 1)


def getFormattedSheet(sheetJson, sheetIndex):
    formatted_josn = json.loads(sheetJson)    
    header = list(formatted_josn["page"][0]["Contractor"][0]["bill_section_items"][0][0].keys())
    number_of_contractors = len(formatted_josn["page"][0]["Contractor"])
    commonColumns = header[0:len(header)-2]
    contractorColums = header[len(header)-2:len(header)]
    header.extend(header[len(header)-2:len(header)]*(number_of_contractors-1))
    startColumn = 'A'
    endColumn = ''
    Row_num_to_insert = 1
    def applyBorderForContractorColumn(nextrow=1):
        for j in range(nextrow):
            for comCol in range(len(commonColumns)):
                sheet[f"{header_column[commonColumns[comCol]]}{Row_num_to_insert+j}"].border = Border(
                    left=Side(border_style="thin",color="000000"),
                    right=Side(border_style="thin", color="000000"))
            for conCol in contractorColums:
                for ConColName in header_column[conCol]:
                    sheet[f"{ConColName}{Row_num_to_insert+j}"].border = Border(
                    right=Side(border_style="thin", color="000000"))
                # sheet[f"{contractor_column_name[i][0]}{Row_num_to_insert+j}"].border = Border(
                #     left=Side(border_style="thin", color="000000"))
                # sheet[f"{contractor_column_name[i][1]}{Row_num_to_insert+j}"].border = Border(
                #     right=Side(border_style="thin", color="000000"))
        
    for i in range(1,len(header)):
        endColumn = increment_column(startColumn if i==1 else endColumn)
    wb_name = "combined.xlsx"
    wb = Workbook() if sheetIndex == 0 else load_workbook(wb_name)
    sheet = wb.create_sheet(formatted_josn["sheet"])
    if sheetIndex == 0:
        del wb["Sheet"] 
    title = str.splitlines(formatted_josn["title"])
    # add titles in the new sheet 
    for i in range(1,len(title)):
        sheet.merge_cells(f"{startColumn}{i}:{endColumn}{i}")
        sheet[f"{startColumn}{i}"]= title[i]
        sheet[f"{startColumn}{i}"].alignment = Alignment(horizontal="center", vertical="center")
        sheet[f"{startColumn}{i}"].font = Font(bold=True)
        Row_num_to_insert = Row_num_to_insert+1
    #creates merged cell contractor name header column
    contracter_merge_cells_start_column = chr(ord(startColumn) + len(commonColumns))
    contracter_merge_cells_end_column = ''
    contractor_column_name = []
    for i in range(1,number_of_contractors+1):
        def contractName():
            startCell =f"{contracter_merge_cells_start_column}{Row_num_to_insert}"
            sheet.merge_cells(f"{startCell}:{contracter_merge_cells_end_column}{Row_num_to_insert}")
            contractor_column_name.append([contracter_merge_cells_start_column,contracter_merge_cells_end_column])
            sheet[startCell]= value_to_insert = formatted_josn["page"][0]["Contractor"][i-1]["contractor_name"]
            sheet[startCell].alignment = Alignment(horizontal="center", vertical="center")
            sheet[startCell].font = Font(bold=True)
            if formatted_josn["sheet"] == "B2-Sec R":
                print("check")
            sheet[startCell].border = apply_border
        if i == 1:
            contracter_merge_cells_end_column = chr(ord(contracter_merge_cells_start_column) + 
                                                    len(contractorColums)-1)
            contractName()
        else:
            contracter_merge_cells_start_column = chr(ord(contracter_merge_cells_end_column) + 1)
            contracter_merge_cells_end_column = chr(ord(contracter_merge_cells_start_column) + len(contractorColums)-1)
            contractName() 
    Row_num_to_insert=Row_num_to_insert+1
    
    commonColumns_item_len = [len(i) for i in commonColumns]
    contractorColumns_item_len = [len(i) for i in contractorColums]
    # getting into each page info
    for page in formatted_josn["page"]:
        # add header to page:
        #page header column names
        print("Writing: ",page["page_name"])
        header_column_count = len(contractorColums)+len(commonColumns)
        header_column = {header[col_inx]: [] if header[col_inx] in contractorColums else "" for col_inx in range(header_column_count)}
        totals_column =[]
        totals_column.extend([[] for _ in range(number_of_contractors)])
        Row_num_to_insert=Row_num_to_insert+1
        header_bg_color = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
        for i in range(1,len(header)+1):
            column_to_insert = f"{chr(ord(startColumn) + (i-1))}" 
            cell_to_insert = f"{column_to_insert}{Row_num_to_insert}"
            sheet[cell_to_insert]=header[i-1]
            sheet[cell_to_insert].font = Font(bold=True)
            sheet[cell_to_insert].alignment = Alignment(horizontal="center", vertical="center")
            sheet[f"{column_to_insert}{Row_num_to_insert-1}"].fill = header_bg_color
            sheet[cell_to_insert].fill = header_bg_color
            sheet[f"{column_to_insert}{Row_num_to_insert+1}"].fill = header_bg_color
            sheet[f"{column_to_insert}{Row_num_to_insert-1}"].border = Border(top=Side(style="thin", color="000000"), 
                                                                              right=Side(style="thin", color="000000"),
                                                                              left=Side(style="thin", color="000000"))
            sheet[cell_to_insert].border = Border(right=Side(style="thin", color="000000"),
                                                  left=Side(style="thin", color="000000"))
            sheet[f"{column_to_insert}{Row_num_to_insert+1}"].border = Border(bottom=Side(style="thin", color="000000"),
                                                                              right=Side(style="thin", color="000000"),
                                                                              left=Side(style="thin", color="000000"))
            if header[i-1] in commonColumns:
                header_column[header[i-1]]= column_to_insert    
            else:
                header_column[header[i-1]].append(column_to_insert)
        Row_num_to_insert=Row_num_to_insert+2
        #page name under "AMOUNT SAR RIYAL" row
        for i in range(1,number_of_contractors+1):
            cell_to_insert = f"{chr(ord(startColumn) + (len(commonColumns)+len(contractorColums*i)-1))}{Row_num_to_insert}"
            sheet[cell_to_insert]=page["page_name"]
            sheet[cell_to_insert].font = Font(bold=True)
            sheet[cell_to_insert].alignment = Alignment(horizontal="center", vertical="center")
        if formatted_josn['page'].index(page) == 0:
            freeze_column = chr(ord(header_column[commonColumns[len(commonColumns)-1]]) + 1)
            sheet.freeze_panes = f"{freeze_column}{Row_num_to_insert}"      
        applyBorderForContractorColumn()
        Row_num_to_insert=Row_num_to_insert+1
        #section_header and section_item
        for section_header_counter in range(0,len(page["Contractor"][0]["bill_section_header"])):
            section_header_split_list = str.splitlines(page["Contractor"][0]["bill_section_header"][section_header_counter])
            header_item_counter = 0
            for section_header_line in section_header_split_list:
                #ignore page_name from section header and empty string
                if page["page_name"] not in section_header_line and len(section_header_line) != 0:
                    cell_to_insert = f"{header_column[f"{header[1]}"]}{Row_num_to_insert}"
                    sheet[cell_to_insert]= section_header_line
                    pre_max_len = commonColumns_item_len[commonColumns.index(header[1])]
                    current_item_len = len(str(section_header_line))
                    if pre_max_len < current_item_len:
                        commonColumns_item_len[commonColumns.index(header[1])] = current_item_len
                    if header_item_counter == 0:
                        sheet[cell_to_insert].font = Font(bold=True)
                    else:
                        sheet[cell_to_insert].font = Font(underline="single")
                    header_item_counter = header_item_counter +1
                    applyBorderForContractorColumn()
                    Row_num_to_insert=Row_num_to_insert+1
                else:
                    continue
            applyBorderForContractorColumn()
            Row_num_to_insert=Row_num_to_insert+1
            for section_item_counter in range(len(page["Contractor"][0]["bill_section_items"][section_header_counter])):
                combained_section_item = {header[col_inx]: [] if header[col_inx] in contractorColums else "" for col_inx in range(header_column_count)}
                # get the items from each contractor combine and store it as dict 
                for contractor in range(0, number_of_contractors):
                    currentContractorSectionHeader = ""
                    currentContractorDescription = ""
                    try:
                        currentContractorSectionHeader= page["Contractor"][contractor]["bill_section_header"][section_header_counter]
                    except IndexError:
                        currentContractorSectionHeader = "section header not found"
                    templateSectionHeader = page["Contractor"][0]["bill_section_header"][section_header_counter]
                    try: 
                        currentContractorDescription = page["Contractor"][contractor]["bill_section_items"][section_header_counter][section_item_counter][header[1]]
                    except IndexError:
                        currentContractorDescription = "section item description not found"
                    templateSectionDescription = page["Contractor"][0]["bill_section_items"][section_header_counter][section_item_counter][header[1]]
                    templateItemQuantity = page["Contractor"][0]["bill_section_items"][section_header_counter][section_item_counter][header[2]]
                    check_header_match = templateSectionHeader == currentContractorSectionHeader
                    check_item_match = currentContractorDescription == templateSectionDescription
                    try:
                        item = page["Contractor"][contractor]["bill_section_items"][section_header_counter][section_item_counter]
                    except IndexError:
                        item = "section description not found"
                    def get_rate_and_amount():
                        if check_header_match and check_item_match and isinstance(item[header[4]],float):
                            rate = float(item[header[4]]) if len(str(item[header[4]]))>0 else "UNPRICED"
                            amount = templateItemQuantity*float(item[header[4]]) if isinstance(rate,float) else "UNPRICED"#templateItemQuantity*rate
                            return {header[4]: rate, header[5]: amount}
                        else:
                            rate= "UNPRICED"
                            amount = "UNPRICED"
                            return {header[4]:rate, header[5]: amount}
                    rate_and_amount = get_rate_and_amount()
                    #assigning values to each item row
                    for header_counter in range(len(commonColumns)+len(contractorColums)):
                        templateSectionItems = page["Contractor"][0]["bill_section_items"][section_header_counter][section_item_counter]
                        if header[header_counter] in commonColumns and contractor == 0:
                            combained_section_item[header[header_counter]] = templateSectionItems[header[header_counter]]
                        elif header[header_counter] in contractorColums:
                            combained_section_item[header[header_counter]].append(rate_and_amount[header[header_counter]])
                # inserting the item values in excel
                for key, value in combained_section_item.items():
                    if key in commonColumns:
                        cell_to_insert = f"{header_column[key]}{Row_num_to_insert}"
                        sheet[cell_to_insert]=value
                        pre_max_len = commonColumns_item_len[commonColumns.index(key)]
                        current_item_len = len(str(value))
                        if pre_max_len < current_item_len:
                            commonColumns_item_len[commonColumns.index(key)] = current_item_len 
                        if key == header[3] or key == header[0]:
                            sheet[cell_to_insert].alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        for i in range(0,len(value)):
                            cell_to_insert = f"{header_column[key][i]}{Row_num_to_insert}" 
                            if key == header[5] and isinstance(value[i],float):
                                sheet[cell_to_insert]=f"=ROUND({header_column[header[2]]}{Row_num_to_insert}*{header_column[header[4]][i]}{Row_num_to_insert},2)"
                                totals_column[i].append(cell_to_insert)
                            else:
                                sheet[cell_to_insert]=round(value[i],2) if isinstance(value[i], (int, float)) else value[i]
                            if value[i]=="UNPRICED":
                                sheet[cell_to_insert].fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
                            elif key == header[4]:
                                #checks all RATE fields are priced
                                if all(isinstance(item, float) for item in value):
                                    if value[i]==max(value):
                                        sheet[cell_to_insert].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
                                else:
                                    priced_value_list=[i for i in value if isinstance(i,float)]
                                    if len(priced_value_list)>1 and value[i]==max(priced_value_list):
                                        sheet[cell_to_insert].fill = PatternFill(start_color="ff0000", end_color="ff0000", fill_type="solid")
                            if not(check_header_match) and check_item_match and value[i]=="UNPRICED":
                                sheet[cell_to_insert].fill = PatternFill(start_color="0000ff", end_color="0000ff", fill_type="solid")
                            val = round(value[i],2) if isinstance(value[i], (int, float)) else value[i]
                            if len(str(val)) > contractorColumns_item_len[contractorColums.index(key)]:
                                contractorColumns_item_len[contractorColums.index(key)] = len(str(val))
                applyBorderForContractorColumn(2)
                Row_num_to_insert = Row_num_to_insert+2                
        for contractor in range(number_of_contractors):
            if contractor == 0:
                startCell = f"{startColumn}{Row_num_to_insert}"
                applyBorderForContractorColumn(2)
                Row_num_to_insert = Row_num_to_insert+2
                endCell = f"{chr(ord(startColumn) + len(commonColumns)-1)}{Row_num_to_insert}"
                sheet.merge_cells(f"{startCell}:{endCell}")
                sheet[startCell]="Total - Saudi Riyal\nCarried to Collection"
                sheet[startCell].alignment = Alignment(horizontal="right", vertical="center",wrap_text=True)
                sheet[startCell].font = Font(bold=True)
                if formatted_josn["sheet"] == "B2-Sec R":
                    print("sum border check")
                sheet[startCell].border = apply_border
            cell_to_insert= f"{header_column[f"{header[5]}"][contractor]}{Row_num_to_insert-1}"
            sheet[cell_to_insert]= f"=ROUND(SUM({','.join(totals_column[contractor])}),2)"
            sheet[cell_to_insert].font = Font(bold=True)
            sheet[f"{header_column[f"{header[5]}"][contractor]}{Row_num_to_insert-2}"].border = Border(
                top=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"))
        applyBorderForContractorColumn()
        for i in range(1,len(header)+1):
            column_to_insert = f"{chr(ord(startColumn) + (i-1))}"
            sheet[f"{column_to_insert}{Row_num_to_insert}"].border = Border(
                bottom=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"))
        sheet.row_breaks.append(Break(id=Row_num_to_insert))
        Row_num_to_insert = Row_num_to_insert+1
    print("commonColumns_item_len: ", commonColumns_item_len)
    print("contractorColumns_item_len: ", contractorColumns_item_len)
    for i in range(len(commonColumns_item_len)):
        sheet.column_dimensions[header_column[commonColumns[i]]].width = commonColumns_item_len[i]+2
    for i in contractorColums:
        for j in header_column[i]:
            sheet.column_dimensions[j].width = contractorColumns_item_len[contractorColums.index(i)]+3
    
    wb.save(wb_name)

    
# getFormattedSheet(formattedSheet)